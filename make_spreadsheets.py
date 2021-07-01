from os import get_terminal_size
import sys
import glob
import subprocess
from pathlib import Path
import pickle
import tqdm

from shutil import copyfile, copytree

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment


main_dir = "/media/hdd/adrien/Friends"
shots_dir = main_dir + "/shots"
videos_dir = main_dir + "/videos"
subvideos_dir = main_dir + "/subvideos"
selected_subvideos_dir = main_dir + "/selected_subvideos"
annotations_dir = main_dir + "/annotations/blank"
problematic_shots_dir = main_dir + "/problematic_shots"

tracks_file = main_dir + "/tracks-features/Friends_final.pk"


def gather_tracks():
    all_tracks = pickle.load(open(tracks_file, "rb"))
    return all_tracks


def gather_shots(episode_number):
    shots_file = "{}/episode{:02d}_shots.txt".format(shots_dir, episode_number)
    shots = []
    with open(shots_file, "r") as f:
        for line in f:
            b, e = tuple(map(int, line.strip().split()))
            shots.append([b, e])
    return shots


def order_tracks(all_tracks, episode_number):
    ep_tracks = all_tracks["episode{:02d}".format(episode_number)]
    body_tracks = ep_tracks['body']
    body_GT = ep_tracks['GT']
    back_tracks = ep_tracks['back']
    back_GT = ep_tracks['back_GT']
    shots = gather_shots(episode_number)

    track_ids_by_shot = [[] for id in range(len(shots))]
    tracks_by_shot = [[] for id in range(len(shots))]
    characters_by_shot = [[] for id in range(len(shots))]

    for track_id in body_tracks:
        track = body_tracks[track_id]
        character = body_GT[track_id]
        
        begin_frame, end_frame = tuple(map(int, track[[0, -1], 0]))
        shot_id = max([id for id, shot in enumerate(shots) if shot[0] <= begin_frame])

        track_ids_by_shot[shot_id].append(track_id)
        tracks_by_shot[shot_id].append(track)
        characters_by_shot[shot_id].append(character)
    
    for track_id in back_tracks:
        track = back_tracks[track_id]
        character = back_GT[track_id]
        
        begin_frame, end_frame = tuple(map(int, track[[0, -1], 0]))
        shot_id = max([id for id, shot in enumerate(shots) if shot[0] < begin_frame])

        track_ids_by_shot[shot_id].append(track_id)
        tracks_by_shot[shot_id].append(track)
        characters_by_shot[shot_id].append(character)
    
    return tracks_by_shot, track_ids_by_shot, characters_by_shot


def have_temporal_intersection(track1, track2, N=1):
    """
    returns true iff the temporal intersection of the two tracks is at least N frame long
    """

    b1, e1 = track1[[0, -1], 0]
    b2, e2 = track2[[0, -1], 0]

    b = max(b1, b2)
    e = min(e1, e2)

    inter = max(0, e - b + 1)

    return inter >= N


def intersect_list(tracks, N=1):
    for i in range(len(tracks)):
        for j in range(i + 1, len(tracks)):
            if have_temporal_intersection(tracks[i], tracks[j], N):
                return True
    return False


def list_problematic_shot_ids(episode_number, all_tracks):
    shots = gather_shots(episode_number)
    tracks_by_shot, track_ids_by_shot, characters_by_shot = order_tracks(all_tracks, episode_number)

    problematic_shot_ids = []
    for shot_id, shot in enumerate(shots):
        tracks = tracks_by_shot[shot_id]
        characters = characters_by_shot[shot_id]
        track_ids = track_ids_by_shot[shot_id]

        begin, end = shot
        length = end - begin + 1
        if length < 50:
            continue
        if len(characters) < 2:
            continue
        if len(characters) != len(list(set(characters))):
            problematic_shot_ids.append(shot_id)

            for character in set(characters):
                character_tracks = []
                character_track_ids = [track_ids[i] for i in range(len(track_ids)) if characters[i] == character]
                for i in range(len(tracks)):
                    if characters[i] == character:
                        character_tracks.append(tracks[i])
                if intersect_list(character_tracks):
                    print("\t", shot_id, character, "\t", character_track_ids)

            # print(shot_id)
            # for i in range(len(tracks)):
            #     character = characters[i]
            #     if characters.count(character) < 2:
            #         continue
            #     track = tracks[i]
            #     print(character, int(track[0,0]), int(track[-1, 0]))
            # print()
    
    return problematic_shot_ids


def generate_problematic_shot_dir(episode_number, all_tracks):
    print("episode {}".format(episode_number))
    problematic_shot_ids = list_problematic_shot_ids(episode_number, all_tracks)
    ep_subvideos_dir = subvideos_dir + "/episode{:02d}".format(episode_number)
    subdir = problematic_shots_dir + "/episode{:02d}".format(episode_number)
    Path(subdir).mkdir(parents=True, exist_ok=True)
    # for shot_id in problematic_shot_ids:
    #     subvideo_file = "{}/episode{:02d}_shot{:03d}.mkv".format(ep_subvideos_dir, episode_number, shot_id)
    #     new_subvideo_file = "{}/episode{:02d}_shot{:03d}.mkv".format(subdir, episode_number, shot_id)
    #     copyfile(subvideo_file, new_subvideo_file)


def generate_all_problematic_shot_dirs():
    all_tracks = gather_tracks()
    for episode_number in tqdm.tqdm(range(1, 26)):
        generate_problematic_shot_dir(episode_number, all_tracks)


def generate_potential_pairs(episode_number, all_tracks):
    shots = gather_shots(episode_number)
    tracks_by_shot, track_ids_by_shot, characters_by_shot = order_tracks(all_tracks, episode_number)

    potential_pairs = []
    for shot_id, shot in enumerate(shots):
        tracks = tracks_by_shot[shot_id]
        characters = characters_by_shot[shot_id]
        track_ids = track_ids_by_shot[shot_id]
        if len(characters) < 2:
            continue
        if len(characters) != len(list(set(characters))):
            skip_shot = False
            for character in set(characters):
                character_tracks = [tracks[i] for i in range(len(tracks)) if characters[i] == character]
                if intersect_list(character_tracks, N=2):
                    skip_shot = True
                    break
            if skip_shot:
                continue

        nb_characters = len(characters)
        for i1 in range(nb_characters):
            for i2 in range(i1 + 1, nb_characters):
                track1 = tracks[i1]
                char1 = characters[i1]
                track_id1 = track_ids[i1]

                track2 = tracks[i2]
                char2 = characters[i2]
                track_id2 = track_ids[i2]

                if not have_temporal_intersection(track1, track2, N=16):
                    continue
                
                char1, char2 = sorted([char1, char2])
                pair = (shot_id, char1, char2)
                potential_pairs.append(pair)

    potential_pairs = list(set(potential_pairs))
    potential_pairs = sorted(potential_pairs)
    potential_pairs = [list(e) for e in potential_pairs]
    return potential_pairs


def generate_spreadsheet(episode_number, all_tracks=None):
    if all_tracks is None:
        all_tracks = gather_tracks()
    potential_pairs = generate_potential_pairs(episode_number, all_tracks)

    Path(annotations_dir).mkdir(parents=True, exist_ok=True)
    excel_file = "{}/annotations_ep{:02d}_blank.xlsx".format(annotations_dir, episode_number)

    workbook = Workbook()
    sheet = workbook.active

    names = ["#", "shot_id", "person1", "person2", "full physical", "partial physical", "full non-physical", "partial non-physical", "no interaction", "undefined"]
    col_letters = dict([(1, "A"), (2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H"), (9, "I"), (10, "J"), (11, "K")])
    for col, name in enumerate(names, start=2):
        sheet.cell(row=2, column=col).value = name
        sheet.cell(row=2, column=col).fill = PatternFill(start_color="00FFFF99", end_color="00FFFF99", fill_type = "solid")
        if col in [4, 5]:
            sheet.column_dimensions[col_letters[col]].width = 20
        elif col > 5:
            sheet.column_dimensions[col_letters[col]].width = len(name)
    for row, pair in enumerate(potential_pairs, start=3):
        entries = [row - 2] + pair
        for col, entry in enumerate(entries, start=2):
            sheet.cell(row=row, column=col).value = entry
            if 3 <= col <= 5:
                sheet.cell(row=row, column=col).fill = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type = "solid")
        for col in range(6, 12):
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    
    workbook.save(filename=excel_file)


def generate_all_spreadsheets():
    all_tracks = gather_tracks()
    for episode_number in tqdm.tqdm(range(1, 26)):
        generate_spreadsheet(episode_number, all_tracks)


def select_and_copy_subvideos(episode_number, all_tracks=None):
    if all_tracks is None:
        all_tracks = gather_tracks()
    potential_pairs = generate_potential_pairs(episode_number, all_tracks)
    shot_ids = []
    for pair in potential_pairs:
        shot_id = pair[0]
        if shot_id not in shot_ids:
            shot_ids.append(shot_id)

    ep_subvideos_dir = subvideos_dir + "/episode{:02d}".format(episode_number)
    subdir = "{}/episode{:02d}".format(selected_subvideos_dir, episode_number)
    Path(subdir).mkdir(parents=True, exist_ok=True)

    for shot_id in shot_ids:
        subvideo_file = "{}/episode{:02d}_shot{:03d}.mkv".format(ep_subvideos_dir, episode_number, shot_id)
        new_subvideo_file = "{}/episode{:02d}_shot{:03d}.mkv".format(subdir, episode_number, shot_id)
        copyfile(subvideo_file, new_subvideo_file)


def select_and_copy_all_subvideos():
    all_tracks = gather_tracks()
    for episode_number in tqdm.tqdm(range(1, 26)):
        select_and_copy_subvideos(episode_number, all_tracks)


def print_out_statistics(all_tracks=None):
    if all_tracks is None:
        all_tracks = gather_tracks()

    print("ep \t shots multi \t problematic shots")
    for episode_number in range(1, 26):
        shots = gather_shots(episode_number)
        tracks_by_shot, track_ids_by_shot, characters_by_shot = order_tracks(all_tracks, episode_number)

        nb_several_person_shots = len([e for e in characters_by_shot if len(list(set(e))) > 1])
        problematic_shot_ids = []
        for shot_id, characters in enumerate(characters_by_shot):
            if len(characters) != len(list(set(characters))):
                problematic_shot_ids.append(shot_id)
        print("{} \t {} \t {}".format(episode_number, nb_several_person_shots, len(problematic_shot_ids)))


if __name__ == "__main__":
    action = sys.argv[1]
    if action == "prob":
        if len(sys.argv) > 2:
            episode_number = int(sys.argv[2])
            generate_problematic_shot_dir(episode_number)
        else:
            generate_all_problematic_shot_dirs()
    elif action == "select":
        if len(sys.argv) > 2:
            episode_number = int(sys.argv[2])
            select_and_copy_subvideos(episode_number)
        else:
            select_and_copy_all_subvideos()
    elif action == "sheets":
        if len(sys.argv) > 2:
            episode_number = int(sys.argv[2])
            generate_spreadsheet(episode_number)
        else:
            generate_all_spreadsheets()
    else:
        print("First argument not recognized")
